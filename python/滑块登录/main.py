import os
import threading
import tkinter as tk
from datetime import datetime
from tkinter import filedialog, messagebox, scrolledtext

import openpyxl

from login_handler import SmartEduLogin


class SmartEduGUI:
    def __init__(self, root: tk.Tk) -> None:
        self.root = root
        self.root.title("SmartEdu 批量登录助手")
        self.root.geometry("640x720")

        # Excel 文件路径
        self.excel_file: str | None = None

        # 积分数据存储
        self.points_data: list[dict] = []

        # 先构建界面，再初始化登录处理器（避免日志回调时控件尚未创建）
        self._build_widgets()

        # 登录处理器
        self.login_handler = SmartEduLogin(self.log_message)

    def _build_widgets(self) -> None:
        # 批量导入区域
        title_batch = tk.Label(self.root, text="批量账号登录（Excel）", font=("微软雅黑", 12, "bold"))
        title_batch.pack(pady=10)

        btn_choose_excel = tk.Button(
            self.root,
            text="选择 Excel 文件（第一行表头，从第2行开始：第一列账号，第二列密码）",
            command=self.on_choose_excel,
            bg="#388E3C",
            fg="white",
        )
        btn_choose_excel.pack(pady=5)

        self.lbl_excel = tk.Label(self.root, text="未选择文件", fg="red")
        self.lbl_excel.pack()

        btn_batch_login = tk.Button(
            self.root,
            text="开始批量登录",
            command=self.on_batch_login,
            bg="#F57C00",
            fg="white",
        )
        btn_batch_login.pack(pady=5)

        # 日志区域
        lbl_log = tk.Label(self.root, text="执行日志", font=("微软雅黑", 12, "bold"))
        lbl_log.pack(pady=10)

        self.txt_log = scrolledtext.ScrolledText(self.root, width=80, height=20)
        self.txt_log.pack(padx=10, pady=5, fill="both", expand=True)

    # ==== 事件处理 ====
    def log_message(self, message: str) -> None:
        self.txt_log.insert(tk.END, message + "\n")
        self.txt_log.see(tk.END)
        # 允许在后台线程中刷新 UI
        self.root.update_idletasks()

    def on_choose_excel(self) -> None:
        path = filedialog.askopenfilename(
            title="选择账号 Excel 文件",
            filetypes=[("Excel 文件", "*.xlsx *.xls")],
        )
        if not path:
            return
        self.excel_file = path
        self.lbl_excel.config(text=f"已选择：{path}", fg="green")
        self.log_message(f"[信息] 已选择 Excel 文件：{path}")

    def on_batch_login(self) -> None:
        """点击“开始批量登录”按钮。"""
        if not self.excel_file:
            messagebox.showwarning("提示", "请先选择包含账号密码的 Excel 文件")
            return

        # 后台线程执行，避免阻塞界面
        t = threading.Thread(target=self._run_batch_login, daemon=True)
        t.start()

    def _run_batch_login(self) -> None:
        """批量登录任务（在后台线程中执行）。"""
        if not self.excel_file:
            self.log_message("[批量-错误] 未选择 Excel 文件")
            return

        try:
            # 清空之前的积分数据
            self.points_data = []
            
            self.log_message(f"[批量-开始] 从 Excel 读取账号：{self.excel_file}")

            wb = openpyxl.load_workbook(self.excel_file)
            sheet = wb.active

            total = 0
            success_count = 0
            fail_count = 0

            for row in sheet.iter_rows(min_row=2, values_only=True):
                # 第一列账号，第二列密码（先转换为字符串，再去除空格）
                account = str(row[0] or "").strip() if row and row[0] is not None else ""
                password = str(row[1] or "").strip() if row and len(row) > 1 and row[1] is not None else ""

                if not account or not password:
                    continue

                total += 1
                self.log_message(f"[批量-开始] 账号：{account}")

                try:
                    ok, msg, points = self.login_handler.login(account, password)
                    if ok:
                        success_count += 1
                        # 记录积分信息
                        points_info = {
                            "account": account,
                            "points": points if points is not None else "获取失败",
                            "fetch_time": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                        }
                        self.points_data.append(points_info)
                        
                        points_str = f"{points}" if points is not None else "获取失败"
                        self.log_message(f"[批量-成功] {account} -> {msg}, 积分: {points_str}")
                    else:
                        fail_count += 1
                        self.log_message(f"[批量-失败] {account} -> {msg}")
                except Exception as exc:  # noqa: BLE001
                    fail_count += 1
                    self.log_message(f"[批量-异常] 账号 {account} -> {exc}")

            if total == 0:
                messagebox.showwarning("提示", "Excel 中没有读取到有效的账号和密码（前两列）。")
                self.log_message("[批量-结束] Excel 中没有有效账号")
                return

            summary = f"批量登录完成：共 {total} 个账号，成功 {success_count} 个，失败 {fail_count} 个"
            self.log_message(f"[批量-汇总] {summary}")
            
            # 导出积分数据
            if len(self.points_data) > 0:
                self.log_message(f"[导出] 开始导出 {len(self.points_data)} 个账号的积分数据...")
                self._export_points_to_excel()
            else:
                self.log_message("[导出] 没有成功获取积分的账号，跳过导出")
            
            messagebox.showinfo("批量登录结果", summary)

        except Exception as exc:  # noqa: BLE001
            self.log_message(f"[批量-异常] {exc}")
            messagebox.showerror("批量登录异常", str(exc))

    def _export_points_to_excel(self) -> None:
        """将积分数据导出为Excel文件。"""
        try:
            # 创建新的工作簿
            wb = openpyxl.Workbook()
            sheet = wb.active
            sheet.title = "积分数据"

            # 添加表头
            headers = ["账号", "积分", "获取时间"]
            for col_idx, header in enumerate(headers, start=1):
                cell = sheet.cell(row=1, column=col_idx, value=header)
                # 设置表头样式（加粗）
                cell.font = openpyxl.styles.Font(bold=True)

            # 写入数据
            for row_idx, data in enumerate(self.points_data, start=2):
                sheet.cell(row=row_idx, column=1, value=data["account"])
                sheet.cell(row=row_idx, column=2, value=data["points"])
                sheet.cell(row=row_idx, column=3, value=data["fetch_time"])

            # 自动调整列宽
            for column in sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except Exception:  # noqa: BLE001
                        pass
                adjusted_width = min(max_length + 2, 50)
                sheet.column_dimensions[column_letter].width = adjusted_width

            # 生成文件名（带时间戳）
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"积分数据_{timestamp}.xlsx"

            # 保存到桌面（如果可以），否则保存到当前目录
            desktop = os.path.join(os.path.expanduser("~"), "Desktop")
            if os.path.exists(desktop):
                filepath = os.path.join(desktop, filename)
            else:
                filepath = filename

            wb.save(filepath)
            self.log_message(f"[导出-成功] 积分数据已导出到: {filepath}")
            messagebox.showinfo("导出成功", f"积分数据已导出到:\n{filepath}")

        except Exception as exc:  # noqa: BLE001
            self.log_message(f"[导出-异常] {exc}")
            messagebox.showerror("导出失败", f"导出积分数据时出错:\n{exc}")


def main() -> None:
    root = tk.Tk()
    app = SmartEduGUI(root)
    root.mainloop()


if __name__ == "__main__":
    main()


